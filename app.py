from flask import Flask, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import io, json, base64
from datetime import datetime, timedelta

app = Flask(__name__)

# ── Shared styles ──
DARK_BLUE = "1F3864"
MED_BLUE = "2E75B6"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
DARK_GRAY = "595959"
RED = "C00000"
GREEN = "00B050"
ORANGE = "ED7D31"

title_font = Font(name="Arial", size=14, bold=True, color=WHITE)
subtitle_font = Font(name="Arial", size=10, color=WHITE)
header_font = Font(name="Arial", size=9, bold=True, color=WHITE)
data_font = Font(name="Arial", size=9)
bold_font = Font(name="Arial", size=9, bold=True)
big_num_font = Font(name="Arial", size=12, bold=True, color=DARK_BLUE)
note_font = Font(name="Arial", size=8, color=DARK_GRAY)
section_font = Font(name="Arial", size=10, bold=True, color=WHITE)

title_fill = PatternFill("solid", fgColor=DARK_BLUE)
subtitle_fill = PatternFill("solid", fgColor=MED_BLUE)
header_fill = PatternFill("solid", fgColor=DARK_BLUE)
light_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
white_fill = PatternFill("solid", fgColor=WHITE)
red_fill = PatternFill("solid", fgColor="FFC7CE")
green_fill = PatternFill("solid", fgColor="C6EFCE")

center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

MARKETPLACES = [
    "Walmart US", "Amazon", "Build with Rise", "Houzz",
    "kbauthority.com", "Lowe's", "Menards", "Overstock",
    "Warehouse USA", "Wayfair COM"
]

EXCLUDE_TYPES = [
    "Accessory Component", "Sinks Component", "Box Component",
    "Other Spare Part", "Sinks Porcelain Component", "Others"
]


def write_title(ws, title, subtitle, max_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    c = ws.cell(row=1, column=1, value=title)
    c.font = title_font; c.fill = title_fill; c.alignment = left
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font = subtitle_font; c.fill = subtitle_fill; c.alignment = left
    ws.row_dimensions[2].height = 22


def write_headers(ws, row, headers, fill=None):
    f = fill or header_fill
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = header_font; c.fill = f; c.alignment = center; c.border = thin_border


def write_data_row(ws, row, values, is_alt=False):
    fill = light_fill if is_alt else white_fill
    for ci, v in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = data_font; c.fill = fill; c.alignment = left; c.border = thin_border


def auto_width(ws, max_col, min_w=10, max_w=30):
    for ci in range(1, max_col + 1):
        col_letter = get_column_letter(ci)
        max_len = min_w
        for row in ws.iter_rows(min_col=ci, max_col=ci, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, max_w))
        ws.column_dimensions[col_letter].width = max_len


def safe_float(v, default=0):
    try: return float(v) if v not in (None, "", "N/A") else default
    except: return default


def detect_col(columns, candidates):
    for c in candidates:
        for col in columns:
            if c.lower() == str(col).lower().strip():
                return col
    return None


def build_price_lookup(master_df, promo_df, pricelist_df):
    sku_col_m = detect_col(master_df.columns, ["SKU", "Item", "Product/Service", "Product Code"])
    if sku_col_m is None:
        sku_col_m = master_df.columns[0]

    lookup = {}
    for _, row in master_df.iterrows():
        sku = str(row[sku_col_m]).strip()
        if not sku or sku == "nan":
            continue
        lookup[sku] = {"retail": None, "retail_src": None, "wc": None, "wc_src": None,
                       "category": str(row.get("Type", row.get("Category", ""))),
                       "name": str(row.get("Description", row.get("Product Name", sku)))}

    if promo_df is not None and len(promo_df) > 0:
        p_sku = detect_col(promo_df.columns, ["SKU", "Item", "Model", "Product Code", "Product SKU"])
        p_retail = detect_col(promo_df.columns, ["Retail", "Retail Price", "List Price", "MSRP", "Retail_US"])
        p_wc = detect_col(promo_df.columns, ["WC", "Wholesale Cost", "Wholesale Price", "Cost", "WC Price"])
        if p_sku:
            for _, row in promo_df.iterrows():
                sku = str(row[p_sku]).strip()
                if sku in lookup:
                    if p_retail and pd.notna(row.get(p_retail)):
                        lookup[sku]["retail"] = safe_float(row[p_retail])
                        lookup[sku]["retail_src"] = "Promo File"
                    if p_wc and pd.notna(row.get(p_wc)):
                        lookup[sku]["wc"] = safe_float(row[p_wc])
                        lookup[sku]["wc_src"] = "Promo File"

    if pricelist_df is not None and len(pricelist_df) > 0:
        pl_sku = detect_col(pricelist_df.columns, ["SKU", "Item", "Model", "Product Code", "Product SKU"])
        pl_retail = detect_col(pricelist_df.columns, ["Retail", "Retail Price", "List Price", "MSRP", "Retail_US"])
        pl_wc = detect_col(pricelist_df.columns, ["WC", "Wholesale Cost", "Wholesale Price", "Cost", "WC Price"])
        if pl_sku:
            for _, row in pricelist_df.iterrows():
                sku = str(row[pl_sku]).strip()
                if sku in lookup:
                    if lookup[sku]["retail"] is None and pl_retail and pd.notna(row.get(pl_retail)):
                        lookup[sku]["retail"] = safe_float(row[pl_retail])
                        lookup[sku]["retail_src"] = "Price List File"
                    if lookup[sku]["wc"] is None and pl_wc and pd.notna(row.get(pl_wc)):
                        lookup[sku]["wc"] = safe_float(row[pl_wc])
                        lookup[sku]["wc_src"] = "Price List File"

    return lookup


def generate_excel(p2s_csv, sales_csv, inventory_csv, master_csv=None, promo_csv=None, pricelist_csv=None):
    p2s_df = pd.read_csv(io.StringIO(p2s_csv)) if p2s_csv else pd.DataFrame()
    sales_df = pd.read_csv(io.StringIO(sales_csv)) if sales_csv else pd.DataFrame()
    inv_df = pd.read_csv(io.StringIO(inventory_csv)) if inventory_csv else pd.DataFrame()
    master_df = pd.read_csv(io.StringIO(master_csv)) if master_csv else None
    promo_df = pd.read_csv(io.StringIO(promo_csv)) if promo_csv else None
    pricelist_df = pd.read_csv(io.StringIO(pricelist_csv)) if pricelist_csv else None

    today = datetime.now()
    date_str = today.strftime("%B %d, %Y")
    d90_ago = today - timedelta(days=90)
    period_str = f"{d90_ago.strftime('%b %d, %Y')} – {today.strftime('%b %d, %Y')}"

    if master_df is not None:
        lookup = build_price_lookup(master_df, promo_df, pricelist_df)
    else:
        lookup = {}

    # ── Detect P2S columns ──
    p2s_sku_col = detect_col(p2s_df.columns, ["SKU", "Item", "Product", "Product Code", "Model"]) or p2s_df.columns[0] if len(p2s_df.columns) > 0 else "SKU"
    p2s_mp_col = detect_col(p2s_df.columns, ["Marketplace", "Channel", "Store", "Retailer", "Shop"]) or (p2s_df.columns[1] if len(p2s_df.columns) > 1 else "Marketplace")
    p2s_price_col = detect_col(p2s_df.columns, ["Price", "Current Price", "Selling Price", "P2S Price", "Marketplace Price"]) or (p2s_df.columns[2] if len(p2s_df.columns) > 2 else "Price")

    # ── Detect Sales columns ──
    s_sku_col = detect_col(sales_df.columns, ["SKU", "Item", "Product/Service", "Product Code"]) or (sales_df.columns[0] if len(sales_df.columns) > 0 else "SKU")
    s_qty_col = detect_col(sales_df.columns, ["Qty", "Quantity", "Units", "Qty."]) or "Qty"
    s_amount_col = detect_col(sales_df.columns, ["Amount", "Total", "Revenue", "Sales Amount", "Selling Price"]) or "Amount"
    s_date_col = detect_col(sales_df.columns, ["Date", "Invoice Date", "Transaction Date"]) or "Date"
    s_mp_col = detect_col(sales_df.columns, ["Marketplace", "Channel", "Class", "Customer"]) or "Class"
    s_inv_col = detect_col(sales_df.columns, ["Invoice", "Invoice Number", "Num", "Invoice No"]) or "Num"

    # ── Detect Inventory columns ──
    i_sku_col = detect_col(inv_df.columns, ["SKU", "Item", "Product", "Product Code"]) or (inv_df.columns[0] if len(inv_df.columns) > 0 else "SKU")
    i_qty_col = detect_col(inv_df.columns, ["Qty", "Available", "Quantity", "Stock", "On Hand", "Qty On Hand"]) or "Available"

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════
    # SHEET: KPI Summary
    # ══════════════════════════════════════════
    ws_kpi = wb.active
    ws_kpi.title = "KPI Summary"
    kpi_headers = ["Marketplace", "Total SKUs Analyzed", "SKUs Listed (Available)", "SKUs Not Listed",
                   "SKUs Out of Stock", "Listing Coverage %", "Below Retail Price", "Above Retail Price",
                   "Aligned", "Alignment %", "Sales Units 90D", "Revenue 90D ($)",
                   "SKUs with No Sales", "Dead Inventory SKUs", "OOS Mismatch", "Unlisted Inventory",
                   "Invoices Reviewed", "Invoices with Violations", "Total Invoice Loss ($)"]

    write_title(ws_kpi, "Marketplace KPI Summary",
                f"All metrics per marketplace | Report Date: {date_str}", len(kpi_headers))
    write_headers(ws_kpi, 3, kpi_headers)

    all_mps = sorted(p2s_df[p2s_mp_col].dropna().unique()) if p2s_mp_col in p2s_df.columns else MARKETPLACES
    kpi_data = []

    for mp_idx, mp in enumerate(all_mps):
        mp_p2s = p2s_df[p2s_df[p2s_mp_col] == mp] if p2s_mp_col in p2s_df.columns else pd.DataFrame()
        total_skus = len(mp_p2s)
        listed = len(mp_p2s[pd.to_numeric(mp_p2s[p2s_price_col], errors="coerce") > 0]) if p2s_price_col in mp_p2s.columns else 0
        oos = len(mp_p2s[pd.to_numeric(mp_p2s[p2s_price_col], errors="coerce") == 0]) if p2s_price_col in mp_p2s.columns else 0
        not_listed = total_skus - listed - oos

        aligned = below = above = 0
        for _, row in mp_p2s.iterrows():
            sku = str(row.get(p2s_sku_col, ""))
            if "-MVP" in sku:
                continue
            price = safe_float(row.get(p2s_price_col))
            if price <= 0:
                continue
            retail = lookup.get(sku, {}).get("retail")
            if retail and retail > 0:
                diff = price - retail
                if diff < -10: below += 1
                elif diff > 10: above += 1
                else: aligned += 1

        evaluated = below + above + aligned
        align_pct = round(aligned / evaluated * 100, 1) if evaluated > 0 else 0

        mp_sales = sales_df[sales_df[s_mp_col].astype(str).str.contains(mp, case=False, na=False)] if s_mp_col in sales_df.columns else pd.DataFrame()
        units_90d = int(pd.to_numeric(mp_sales[s_qty_col], errors="coerce").sum()) if s_qty_col in mp_sales.columns else 0
        rev_90d = round(pd.to_numeric(mp_sales[s_amount_col], errors="coerce").sum(), 2) if s_amount_col in mp_sales.columns else 0
        inv_reviewed = len(mp_sales)
        violations = 0
        total_loss = 0

        no_sales = total_skus - len(mp_sales[s_sku_col].unique()) if s_sku_col in mp_sales.columns and total_skus > 0 else total_skus

        row_data = [mp, total_skus, listed, not_listed, oos, round(listed/total_skus*100,1) if total_skus else 0,
                    below, above, aligned, align_pct, units_90d, rev_90d, no_sales, 0, 0, 0,
                    inv_reviewed, violations, total_loss]
        kpi_data.append(row_data)
        write_data_row(ws_kpi, 4 + mp_idx, row_data, mp_idx % 2 == 1)

    auto_width(ws_kpi, len(kpi_headers))

    # ══════════════════════════════════════════
    # SHEET: Price Lookup Table
    # ══════════════════════════════════════════
    ws_pl = wb.create_sheet("Price Lookup Table")
    pl_headers = ["SKU", "Retail Price ($)", "Retail Source", "WC ($)", "WC Source"]
    write_title(ws_pl, "Price Lookup Table",
                "Retail & WC prices per SKU | Source: Promo File (priority) → Price List File", len(pl_headers))
    write_headers(ws_pl, 3, pl_headers)
    ri = 4
    for sku, info in sorted(lookup.items()):
        write_data_row(ws_pl, ri, [sku, info["retail"], info["retail_src"], info["wc"], info["wc_src"]], (ri-4) % 2 == 1)
        ri += 1
    auto_width(ws_pl, len(pl_headers))

    # ══════════════════════════════════════════
    # SHEET: Price Alignment — All MPs
    # ══════════════════════════════════════════
    ws_pa = wb.create_sheet("Price Alignment — All MPs")
    pa_headers = ["SKU", "Category", "Product Name", "Marketplace", "P2S Price ($)",
                  "Retail Price ($)", "Retail Source", "WC ($)", "WC Source",
                  "Price Diff ($)", "Pricing Status", "Below Cost Risk"]

    total_eval = 0; total_aligned = 0; total_below = 0; total_above = 0; total_bcr = 0
    pa_rows = []
    for _, row in p2s_df.iterrows():
        sku = str(row.get(p2s_sku_col, ""))
        if "-MVP" in sku: continue
        price = safe_float(row.get(p2s_price_col))
        if price <= 0: continue
        mp = str(row.get(p2s_mp_col, ""))
        info = lookup.get(sku, {})
        retail = info.get("retail")
        wc = info.get("wc")
        if not retail or retail <= 0: continue
        diff = round(price - retail, 2)
        if diff < -10: status = "Below Retail Price"; total_below += 1
        elif diff > 10: status = "Above Retail Price"; total_above += 1
        else: status = "Aligned"; total_aligned += 1
        total_eval += 1
        bcr = "Yes" if wc and price < wc else None
        if bcr: total_bcr += 1
        pa_rows.append([sku, info.get("category",""), info.get("name",""), mp, price,
                        retail, info.get("retail_src",""), wc, info.get("wc_src",""),
                        diff, status, bcr])

    align_pct_all = round(total_aligned/total_eval*100, 1) if total_eval else 0
    write_title(ws_pa, "Price Alignment — All Marketplaces (Available SKUs, Excl. -MVP)",
                f"Total Evaluated: {total_eval} | Aligned: {total_aligned} ({align_pct_all}%) | Below MAP: {total_below} | Above MAP: {total_above} | Below Cost Risk: {total_bcr}",
                len(pa_headers))

    ws_pa.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(pa_headers))
    c = ws_pa.cell(row=3, column=1, value="ALIGNMENT SUMMARY BY MARKETPLACE (sorted: worst first)")
    c.font = section_font; c.fill = subtitle_fill; c.alignment = left

    write_headers(ws_pa, 4, pa_headers)
    for ri, pa_row in enumerate(pa_rows):
        write_data_row(ws_pa, 5 + ri, pa_row, ri % 2 == 1)
    auto_width(ws_pa, len(pa_headers))

    # ══════════════════════════════════════════
    # SHEETS: Per-Marketplace Detail
    # ══════════════════════════════════════════
    mp_headers = ["SKU", "Category", "Product Name", "Marketplace", "P2S Price ($)",
                  "Retail Price ($)", "Retail Source", "WC ($)", "WC Source",
                  "Price Diff ($)", "Pricing Status", "Below Cost Risk"]

    for mp in all_mps:
        safe_name = mp.replace("/", "_").replace("\\", "_").replace(":", "")[:31]
        ws_mp = wb.create_sheet(safe_name)
        mp_p2s = p2s_df[p2s_df[p2s_mp_col] == mp] if p2s_mp_col in p2s_df.columns else pd.DataFrame()
        listed = len(mp_p2s[pd.to_numeric(mp_p2s[p2s_price_col], errors="coerce") > 0]) if p2s_price_col in mp_p2s.columns else 0
        oos = len(mp_p2s[pd.to_numeric(mp_p2s[p2s_price_col], errors="coerce") == 0]) if p2s_price_col in mp_p2s.columns else 0
        not_listed = len(mp_p2s) - listed - oos

        write_title(ws_mp, f"📊 {mp} — Marketplace Detail",
                    f"Listed: {listed} | OOS: {oos} | Not Listed: {not_listed}", len(mp_headers))

        ws_mp.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(mp_headers))
        c = ws_mp.cell(row=3, column=1, value="A. PRICE ALIGNMENT (Available SKUs, excluding -MVP)")
        c.font = section_font; c.fill = subtitle_fill; c.alignment = left

        write_headers(ws_mp, 4, mp_headers)
        ri = 5
        for _, row in mp_p2s.iterrows():
            sku = str(row.get(p2s_sku_col, ""))
            if "-MVP" in sku: continue
            price = safe_float(row.get(p2s_price_col))
            if price <= 0: continue
            info = lookup.get(sku, {})
            retail = info.get("retail")
            wc = info.get("wc")
            if not retail or retail <= 0: continue
            diff = round(price - retail, 2)
            if diff < -10: status = "Below Retail Price"
            elif diff > 10: status = "Above Retail Price"
            else: status = "Aligned"
            bcr = "Yes" if wc and price < wc else None
            write_data_row(ws_mp, ri, [sku, info.get("category",""), info.get("name",""), mp, price,
                                        retail, info.get("retail_src",""), wc, info.get("wc_src",""),
                                        diff, status, bcr], (ri-5) % 2 == 1)
            ri += 1
        auto_width(ws_mp, len(mp_headers))

    # ══════════════════════════════════════════
    # SHEET: Listing Coverage
    # ══════════════════════════════════════════
    ws_lc = wb.create_sheet("Listing Coverage")
    lc_headers = ["SKU", "Category", "Product Name", "Marketplace", "P2S Price ($)", "Listing_Status"]
    total_avail = total_oos = total_nl = 0
    lc_rows = []
    for _, row in p2s_df.iterrows():
        sku = str(row.get(p2s_sku_col, ""))
        mp = str(row.get(p2s_mp_col, ""))
        price = row.get(p2s_price_col)
        price_num = safe_float(price)
        if pd.isna(price) or str(price).strip() == "":
            status = "Not Listed"; total_nl += 1
        elif price_num == 0:
            status = "Out of Stock"; total_oos += 1
        else:
            status = "Available"; total_avail += 1
        info = lookup.get(sku, {})
        lc_rows.append([sku, info.get("category",""), info.get("name",""), mp, price_num if status != "Not Listed" else "", status])

    write_title(ws_lc, "Listing Coverage — All SKUs × All Marketplaces",
                f"Available: {total_avail:,} | OOS: {total_oos:,} | Not Listed: {total_nl:,}", len(lc_headers))
    write_headers(ws_lc, 3, lc_headers)
    for ri, lc_row in enumerate(lc_rows):
        write_data_row(ws_lc, 4 + ri, lc_row, ri % 2 == 1)
    auto_width(ws_lc, len(lc_headers))

    # ══════════════════════════════════════════
    # SHEET: OOS vs Inventory Mismatch
    # ══════════════════════════════════════════
    ws_oos = wb.create_sheet("OOS vs Inventory Mismatch")
    oos_headers = ["SKU", "Category", "Product Name", "Marketplace", "P2S Price ($)",
                   "Listing Status", "USA Inventory", "Inventory_Mismatch_Flag"]
    oos_rows = []
    inv_dict = {}
    if len(inv_df) > 0 and i_sku_col in inv_df.columns:
        for _, row in inv_df.iterrows():
            sku = str(row.get(i_sku_col, "")).strip()
            qty = safe_float(row.get(i_qty_col))
            inv_dict[sku] = inv_dict.get(sku, 0) + qty

    oos_count = unlisted_count = 0
    for _, row in p2s_df.iterrows():
        sku = str(row.get(p2s_sku_col, ""))
        mp = str(row.get(p2s_mp_col, ""))
        price = row.get(p2s_price_col)
        price_num = safe_float(price)
        inv_qty = inv_dict.get(sku, 0)
        if inv_qty <= 0: continue
        if pd.isna(price) or str(price).strip() == "":
            flag = "Unlisted Inventory"; unlisted_count += 1; ls = "Not Listed"
        elif price_num == 0:
            flag = "OOS Mismatch"; oos_count += 1; ls = "Out of Stock"
        else:
            continue
        info = lookup.get(sku, {})
        oos_rows.append([sku, info.get("category",""), info.get("name",""), mp, price_num, ls, inv_qty, flag])

    write_title(ws_oos, "Out of Stock vs Inventory Mismatch — All Marketplaces",
                f"OOS Mismatch: {oos_count} | Unlisted Inventory: {unlisted_count}", len(oos_headers))
    write_headers(ws_oos, 3, oos_headers)
    for ri, r in enumerate(oos_rows):
        write_data_row(ws_oos, 4 + ri, r, ri % 2 == 1)
    auto_width(ws_oos, len(oos_headers))

    # ══════════════════════════════════════════
    # SHEET: Invoice Violations
    # ══════════════════════════════════════════
    ws_inv = wb.create_sheet("Invoice Violations")
    inv_headers = ["Invoice Date", "Invoice Number", "SKU", "Marketplace", "Selling Price",
                   "Allowed Invoice Price", "Price Source", "WC", "WC Source",
                   "Difference", "Difference vs WC", "Quantity", "Violation",
                   "WC Violation", "Total Loss", "Total Loss vs WC"]
    inv_rows = []
    total_violations = 0; total_loss_val = 0; wc_violations = 0; wc_loss_val = 0

    for _, row in sales_df.iterrows():
        sku = str(row.get(s_sku_col, "")).strip()
        info = lookup.get(sku, {})
        retail = info.get("retail")
        wc = info.get("wc")
        sell_price = safe_float(row.get(s_amount_col))
        qty = safe_float(row.get(s_qty_col), 1)
        if qty == 0: qty = 1
        unit_price = sell_price / qty if qty else sell_price
        date = row.get(s_date_col, "")
        inv_num = row.get(s_inv_col, "")
        mp = str(row.get(s_mp_col, ""))

        violation = "No Violation"
        loss = 0
        wc_viol = "No"
        wc_loss = 0
        diff = 0; diff_wc = 0

        if retail and retail > 0:
            diff = round(unit_price - retail, 2)
            if unit_price < retail:
                violation = "Invoice Violation"
                loss = round((retail - unit_price) * qty, 2)
                total_violations += 1
                total_loss_val += loss

        if wc and wc > 0:
            diff_wc = round(unit_price - wc, 2)
            if unit_price < wc:
                wc_viol = "Below WC"
                wc_loss = round((wc - unit_price) * qty, 2)
                wc_violations += 1
                wc_loss_val += wc_loss

        inv_rows.append([date, inv_num, sku, mp, unit_price, retail, info.get("retail_src",""),
                         wc, info.get("wc_src",""), diff, diff_wc, qty, violation, wc_viol, loss, wc_loss])

    write_title(ws_inv, "Invoice Violation Analysis",
                f"Violations: {total_violations:,} | Total Loss vs MAP: ${total_loss_val:,.2f} | WC Violations: {wc_violations} | Total Loss vs WC: ${wc_loss_val:,.2f}",
                len(inv_headers))

    ws_inv.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
    c = ws_inv.cell(row=3, column=1, value="SUMMARY BY MARKETPLACE")
    c.font = section_font; c.fill = subtitle_fill; c.alignment = left

    sum_headers = ["Marketplace", "Invoices Reviewed", "Invoice Violations", "Total Loss ($)", "WC Violations", "Total Loss vs WC ($)"]
    write_headers(ws_inv, 4, sum_headers)
    ri = 5
    # Summary by marketplace
    if s_mp_col in sales_df.columns:
        for mp in sorted(sales_df[s_mp_col].dropna().unique()):
            mp_inv = [r for r in inv_rows if r[3] == mp]
            reviewed = len(mp_inv)
            viols = sum(1 for r in mp_inv if r[12] == "Invoice Violation")
            loss = sum(r[14] for r in mp_inv)
            wv = sum(1 for r in mp_inv if r[13] == "Below WC")
            wl = sum(r[15] for r in mp_inv)
            write_data_row(ws_inv, ri, [mp, reviewed, viols, round(loss,2), wv, round(wl,2)], (ri-5) % 2 == 1)
            ri += 1

    ri += 1
    ws_inv.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(inv_headers))
    c = ws_inv.cell(row=ri, column=1, value="DETAILED INVOICE RECORDS")
    c.font = section_font; c.fill = subtitle_fill; c.alignment = left
    ri += 1
    write_headers(ws_inv, ri, inv_headers)
    ri += 1
    for idx, r in enumerate(inv_rows):
        write_data_row(ws_inv, ri, r, idx % 2 == 1)
        ri += 1
    auto_width(ws_inv, len(inv_headers))

    # ══════════════════════════════════════════
    # SHEET: Sales Performance 90D
    # ══════════════════════════════════════════
    ws_sales = wb.create_sheet("Sales Performance 90D")
    sp_headers = ["SKU", "Category", "Product Name", "Marketplace", "Units Sold 90D",
                  "Revenue 90D ($)", "Orders 90D", "Last Sale Date", "Sales_Status"]

    total_rev = pd.to_numeric(sales_df[s_amount_col], errors="coerce").sum() if s_amount_col in sales_df.columns else 0

    write_title(ws_sales, "Sales Performance — Last 90 Days",
                f"Period: {period_str} | Total Revenue: ${total_rev:,.2f}", len(sp_headers))

    ws_sales.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(sp_headers))
    c = ws_sales.cell(row=3, column=1, value="SALES SUMMARY BY MARKETPLACE")
    c.font = section_font; c.fill = subtitle_fill; c.alignment = left

    mp_sum_headers = ["Marketplace", "Units Sold 90D", "Revenue 90D ($)", "Orders 90D", "Unique SKUs Sold"]
    write_headers(ws_sales, 4, mp_sum_headers)
    ri = 5
    if s_mp_col in sales_df.columns:
        for mp in sorted(sales_df[s_mp_col].dropna().unique()):
            mp_s = sales_df[sales_df[s_mp_col] == mp]
            units = int(pd.to_numeric(mp_s[s_qty_col], errors="coerce").sum()) if s_qty_col in mp_s.columns else 0
            rev = round(pd.to_numeric(mp_s[s_amount_col], errors="coerce").sum(), 2) if s_amount_col in mp_s.columns else 0
            orders = len(mp_s)
            unique_skus = mp_s[s_sku_col].nunique() if s_sku_col in mp_s.columns else 0
            write_data_row(ws_sales, ri, [mp, units, rev, orders, unique_skus], (ri-5) % 2 == 1)
            ri += 1

    ri += 1
    ws_sales.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(sp_headers))
    c = ws_sales.cell(row=ri, column=1, value="DETAILED SKU PERFORMANCE")
    c.font = section_font; c.fill = subtitle_fill; c.alignment = left
    ri += 1
    write_headers(ws_sales, ri, sp_headers)
    ri += 1

    if s_sku_col in sales_df.columns:
        sku_group = sales_df.groupby(s_sku_col)
        for sku, group in sku_group:
            info = lookup.get(str(sku), {})
            units = int(pd.to_numeric(group[s_qty_col], errors="coerce").sum()) if s_qty_col in group.columns else 0
            rev = round(pd.to_numeric(group[s_amount_col], errors="coerce").sum(), 2) if s_amount_col in group.columns else 0
            orders = len(group)
            last_date = str(group[s_date_col].max()) if s_date_col in group.columns else ""
            status = "Active" if units > 0 else "No Sales"
            mp_list = ", ".join(group[s_mp_col].dropna().unique()) if s_mp_col in group.columns else ""
            write_data_row(ws_sales, ri, [sku, info.get("category",""), info.get("name",""),
                                          mp_list, units, rev, orders, last_date, status], (ri % 2 == 1))
            ri += 1
    auto_width(ws_sales, len(sp_headers))

    # ══════════════════════════════════════════
    # SHEET: Weekly Sales Trend
    # ══════════════════════════════════════════
    ws_wk = wb.create_sheet("Weekly Sales Trend")
    wk_headers = ["Week (Mon Start)", "Marketplace", "Units Sold", "Revenue ($)", "Orders"]
    write_title(ws_wk, "Weekly Sales Trend",
                f"Week = Mon–Sun | Period: {period_str}", len(wk_headers))
    write_headers(ws_wk, 3, wk_headers)
    # Simplified weekly aggregation
    ri = 4
    if s_date_col in sales_df.columns:
        sales_df["_date"] = pd.to_datetime(sales_df[s_date_col], errors="coerce")
        sales_df["_week"] = sales_df["_date"].dt.to_period("W-SUN").apply(lambda x: x.start_time.strftime("%Y-%m-%d") if pd.notna(x) else "")
        for (wk, mp), group in sales_df.groupby(["_week", s_mp_col] if s_mp_col in sales_df.columns else ["_week"]):
            units = int(pd.to_numeric(group[s_qty_col], errors="coerce").sum()) if s_qty_col in group.columns else 0
            rev = round(pd.to_numeric(group[s_amount_col], errors="coerce").sum(), 2) if s_amount_col in group.columns else 0
            orders = len(group)
            mp_val = mp if isinstance(mp, str) else ""
            write_data_row(ws_wk, ri, [wk, mp_val, units, rev, orders], (ri-4) % 2 == 1)
            ri += 1
    auto_width(ws_wk, len(wk_headers))

    # ══════════════════════════════════════════
    # SHEET: Inventory Risk
    # ══════════════════════════════════════════
    ws_ir = wb.create_sheet("Inventory Risk")
    ir_headers = ["SKU", "Category", "Product Name", "USA Inventory", "Units Sold 90D",
                  "Revenue 90D ($)", "Sales_Status", "Dead_Inventory"]

    sales_by_sku = {}
    if s_sku_col in sales_df.columns:
        for sku, group in sales_df.groupby(s_sku_col):
            sales_by_sku[str(sku)] = {
                "units": int(pd.to_numeric(group[s_qty_col], errors="coerce").sum()) if s_qty_col in group.columns else 0,
                "revenue": round(pd.to_numeric(group[s_amount_col], errors="coerce").sum(), 2) if s_amount_col in group.columns else 0
            }

    total_inv_units = sum(inv_dict.values())
    dead_count = sum(1 for sku, qty in inv_dict.items() if qty > 0 and sales_by_sku.get(sku, {}).get("units", 0) == 0)

    write_title(ws_ir, "Inventory Risk Analysis",
                f"Total Inventory: {int(total_inv_units):,} units | Dead Inventory SKUs: {dead_count}", len(ir_headers))

    ws_ir.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(ir_headers))
    c = ws_ir.cell(row=3, column=1, value="A. DEAD INVENTORY (Inventory > 0, Zero Sales in 90 Days)")
    c.font = section_font; c.fill = subtitle_fill; c.alignment = left

    write_headers(ws_ir, 4, ir_headers)
    ri = 5
    for sku, qty in sorted(inv_dict.items()):
        if qty <= 0: continue
        s = sales_by_sku.get(sku, {"units": 0, "revenue": 0})
        status = "Active" if s["units"] > 0 else "No Sales"
        dead = "Dead Inventory" if s["units"] == 0 else ""
        info = lookup.get(sku, {})
        write_data_row(ws_ir, ri, [sku, info.get("category",""), info.get("name",""),
                                    qty, s["units"], s["revenue"], status, dead], (ri-5) % 2 == 1)
        ri += 1
    auto_width(ws_ir, len(ir_headers))

    # ══════════════════════════════════════════
    # SHEET: Daily Snapshot
    # ══════════════════════════════════════════
    ws_ds = wb.create_sheet("Daily Snapshot")
    ds_headers = ["Date", "Country", "Marketplace", "Total SKUs Listed", "Alignment %",
                  "Below Retail Price", "Above Retail Price", "Not Listed",
                  "Out of Stock", "No Sales 90D", "Invoice Violations"]
    write_title(ws_ds, "Control Tower — Daily Snapshot Log",
                f"Updated: {date_str} | Country: US | Auto-appended daily", len(ds_headers))
    write_headers(ws_ds, 3, ds_headers)
    ri = 4
    date_today = today.strftime("%Y-%m-%d")
    for kpi in kpi_data:
        write_data_row(ws_ds, ri, [date_today, "US", kpi[0], kpi[2], kpi[9],
                                    kpi[6], kpi[7], kpi[3], kpi[4], kpi[12], kpi[17]], (ri-4) % 2 == 1)
        ri += 1
    auto_width(ws_ds, len(ds_headers))

    # ══════════════════════════════════════════
    # SHEET: Executive Summary (first sheet)
    # ══════════════════════════════════════════
    ws_ex = wb.create_sheet("Executive Summary", 0)
    max_c = 18
    write_title(ws_ex, "  MARKETPLACE CONTROL TOWER — EXECUTIVE SUMMARY  |  Stylish USA Inc.",
                f"Report Date: {date_str}  |  90-Day Period: {period_str}  |  Marketplaces: {len(all_mps)}  |  Master Catalog SKUs: {len(lookup)}",
                max_c)

    r = 4
    for ci, sec in [(1, "CATALOG & COVERAGE"), (5, "PRICE ALIGNMENT"), (10, "INVENTORY & LISTING"), (14, "SALES & INVOICES")]:
        end_ci = ci + 3 if ci < 14 else ci + 4
        ws_ex.merge_cells(start_row=r, start_column=ci, end_row=r, end_column=end_ci)
        c = ws_ex.cell(row=r, column=ci, value=sec)
        c.font = header_font; c.fill = title_fill; c.alignment = center

    metrics = [
        (5, 1, "Master Catalog SKUs", str(len(lookup)), "After product type exclusions", 5, "Listings Evaluated", str(total_eval)),
        (6, 1, "Total Marketplaces", str(len(all_mps)), "All P2S marketplaces", 6, "Aligned", f"{total_aligned} ({align_pct_all}%)"),
        (7, 1, "Available Listings", str(total_avail), "P2S price > 0", 7, "Below Retail", str(total_below)),
        (8, 1, "Out of Stock", str(total_oos), "P2S price = 0", 8, "Above Retail", str(total_above)),
    ]
    for row_data in metrics:
        row_num = row_data[0]
        fill = light_fill if row_num % 2 == 0 else white_fill
        ws_ex.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        c = ws_ex.cell(row=row_num, column=1, value=row_data[2])
        c.font = bold_font; c.fill = fill; c.alignment = left
        c = ws_ex.cell(row=row_num, column=3, value=row_data[3])
        c.font = big_num_font; c.fill = fill; c.alignment = center
        c = ws_ex.cell(row=row_num, column=4, value=row_data[4])
        c.font = note_font; c.fill = fill; c.alignment = left

    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


@app.route("/generate", methods=["POST"])
def generate():
    data = request.json
    p2s_csv = data.get("p2s_csv", "")
    sales_csv = data.get("sales_csv", "")
    inventory_csv = data.get("inventory_csv", "")
    master_csv = data.get("master_csv", "")
    promo_csv = data.get("promo_csv", "")
    pricelist_csv = data.get("pricelist_csv", "")

    output = generate_excel(p2s_csv, sales_csv, inventory_csv, master_csv, promo_csv, pricelist_csv)

    if data.get("return_base64"):
        b64 = base64.b64encode(output.read()).decode("utf-8")
        return jsonify({"file": b64, "filename": f"Marketplace_Control_Tower_USA_{datetime.now().strftime('%Y_%m_%d')}.xlsx"})

    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     download_name=f"Marketplace_Control_Tower_USA_{datetime.now().strftime('%Y_%m_%d')}.xlsx")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)
